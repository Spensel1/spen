from openpyxl import Workbook
import datetime


def test_Workbook_Cells():
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['A1'] = 42
    assert ws['A1'].value == 42
    assert isinstance(ws['A1'].value, int)

    # Rows can also be appended
    ws.append([1, 2, 3])

    # Python types will automatically be converted
    date = '18/05/2020 - 18:05:12'

# convert string to datetimeformat
    ws['A2'] = datetime.datetime.strptime(date, "%d/%m/%Y - %H:%M:%S")
    assert str(ws['A2'].value) == '2020-05-18 18:05:12'

if __name__ == "__main__":
    test_Workbook_Cells()