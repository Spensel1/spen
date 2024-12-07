#  tests for openpyxl-3.1.5-py312h827c3e9_0 (this is a generated file);
print('===== testing package: openpyxl-3.1.5-py312h827c3e9_0 =====');
print('running run_test.py');
#  --- run_test.py (begin) ---
from openpyxl import Workbook
import datetime


def test_Workbook_Cells():
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['A1'] = 42
    assert ws['A1'].value == 42
    assert isinstance(ws['A1'].value, int)

    # Rows can also be appended
    ws.append([1, 2, 3])

    # Python types will automatically be converted
    date = '18/05/2020 - 18:05:12'

# convert string to datetimeformat
    ws['A2'] = datetime.datetime.strptime(date, "%d/%m/%Y - %H:%M:%S")
    assert str(ws['A2'].value) == '2020-05-18 18:05:12'

if __name__ == "__main__":
    test_Workbook_Cells()#  --- run_test.py (end) ---

print('===== openpyxl-3.1.5-py312h827c3e9_0 OK =====');
print("import: 'openpyxl'")
import openpyxl

print("import: 'openpyxl.cell'")
import openpyxl.cell

print("import: 'openpyxl.chart'")
import openpyxl.chart

print("import: 'openpyxl.chartsheet'")
import openpyxl.chartsheet

print("import: 'openpyxl.comments'")
import openpyxl.comments

print("import: 'openpyxl.compat'")
import openpyxl.compat

print("import: 'openpyxl.descriptors'")
import openpyxl.descriptors

print("import: 'openpyxl.drawing'")
import openpyxl.drawing

print("import: 'openpyxl.formatting'")
import openpyxl.formatting

print("import: 'openpyxl.formula'")
import openpyxl.formula

print("import: 'openpyxl.packaging'")
import openpyxl.packaging

print("import: 'openpyxl.pivot'")
import openpyxl.pivot

print("import: 'openpyxl.reader'")
import openpyxl.reader

print("import: 'openpyxl.styles'")
import openpyxl.styles

print("import: 'openpyxl.utils'")
import openpyxl.utils

print("import: 'openpyxl.workbook'")
import openpyxl.workbook

print("import: 'openpyxl.workbook.external_link'")
import openpyxl.workbook.external_link

print("import: 'openpyxl.worksheet'")
import openpyxl.worksheet

print("import: 'openpyxl.writer'")
import openpyxl.writer

print("import: 'openpyxl.xml'")
import openpyxl.xml

